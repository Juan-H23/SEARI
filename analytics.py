"""
analytics.py — Módulo de análisis e inteligencia de SEARI
"""

import io
import csv
import json
from datetime import date, datetime, timedelta
from collections import defaultdict
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session

# ReportLab imports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# OpenPyXL imports
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

# CORREGIDO: Import desde database.models

from app.services.db_service import SessionLocal, engine
from app.database.models import Registro, Base  # Necesitas crear SessionLocal en database/__init__.py

router = APIRouter()

# ─── DB dependency ───────────────────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ... resto de tu código analytics.py (las funciones fmt, pct_change, etc.) ...
# ─── Helpers ─────────────────────────────────────────────────────────────────
def fmt(v: float) -> str:
    return f"${v:,.2f}"

def pct_change(a: float, b: float) -> Optional[float]:
    """Variación porcentual de a → b. None si a==0."""
    if a == 0:
        return None
    return ((b - a) / abs(a)) * 100

def periodo_label(periodo: str, key: str) -> str:
    if periodo == "diario":
        return key                          # "2025-01-15"
    if periodo == "mensual":
        y, m = key.split("-")
        meses = ["Ene","Feb","Mar","Abr","May","Jun",
                 "Jul","Ago","Sep","Oct","Nov","Dic"]
        return f"{meses[int(m)-1]} {y}"
    return key                              # anual: "2025"

# ─── Core analytics engine ───────────────────────────────────────────────────
def compute_analytics(registros: list, periodo: str = "mensual") -> dict:
    if not registros:
        return {"empty": True}

    # ── Agrupar por periodo ───────────────────────────────────────────────────
    grupos: dict = defaultdict(lambda: {"ingresos": 0.0, "egresos": 0.0, "n": 0})

    for r in registros:
        f = r.fecha if isinstance(r.fecha, date) else date.fromisoformat(str(r.fecha))
        if periodo == "diario":
            key = f.isoformat()
        elif periodo == "mensual":
            key = f"{f.year}-{f.month:02d}"
        else:
            key = str(f.year)

        if r.tipo.lower() == "ingreso":
            grupos[key]["ingresos"] += float(r.valor)
        else:
            grupos[key]["egresos"]  += float(r.valor)
        grupos[key]["n"] += 1

    sorted_keys = sorted(grupos.keys())
    periodos = []
    for k in sorted_keys:
        g = grupos[k]
        balance = g["ingresos"] - g["egresos"]
        margen  = (balance / g["ingresos"] * 100) if g["ingresos"] > 0 else 0.0
        periodos.append({
            "key":      k,
            "label":    periodo_label(periodo, k),
            "ingresos": round(g["ingresos"], 2),
            "egresos":  round(g["egresos"],  2),
            "balance":  round(balance, 2),
            "margen":   round(margen, 2),
            "n":        g["n"],
        })

    # ── Totales globales ──────────────────────────────────────────────────────
    total_ing = sum(r["ingresos"] for r in periodos)
    total_eg  = sum(r["egresos"]  for r in periodos)
    total_bal = total_ing - total_eg
    total_mrg = (total_bal / total_ing * 100) if total_ing > 0 else 0.0
    n_total   = sum(r["n"] for r in periodos)

    # ── Variaciones periodo a periodo ────────────────────────────────────────
    for i, p in enumerate(periodos):
        if i == 0:
            p["var_ingresos"] = None
            p["var_egresos"]  = None
            p["var_balance"]  = None
        else:
            prev = periodos[i - 1]
            p["var_ingresos"] = pct_change(prev["ingresos"], p["ingresos"])
            p["var_egresos"]  = pct_change(prev["egresos"],  p["egresos"])
            p["var_balance"]  = pct_change(prev["balance"],  p["balance"])

    # ── Detección de valores atípicos (IQR) ──────────────────────────────────
    valores = [float(r.valor) for r in registros]
    valores.sort()
    n = len(valores)
    q1 = valores[n // 4]
    q3 = valores[(3 * n) // 4]
    iqr = q3 - q1
    umbral_alto = q3 + 1.5 * iqr
    umbral_bajo = q1 - 1.5 * iqr

    atipicos = [
        {
            "id":       r.id,
            "concepto": r.concepto,
            "valor":    float(r.valor),
            "fecha":    str(r.fecha),
            "tipo":     r.tipo,
            "razon":    "valor elevado" if float(r.valor) > umbral_alto else "valor muy bajo",
        }
        for r in registros
        if float(r.valor) > umbral_alto or float(r.valor) < umbral_bajo
    ]

    # ── Mejor / peor periodo ─────────────────────────────────────────────────
    mejor  = max(periodos, key=lambda p: p["balance"])
    peor   = min(periodos, key=lambda p: p["balance"])
    max_ing = max(periodos, key=lambda p: p["ingresos"])
    max_eg  = max(periodos, key=lambda p: p["egresos"])

    # ── Tendencia (regresión lineal simple sobre balances) ───────────────────
    if len(periodos) >= 3:
        xs = list(range(len(periodos)))
        ys = [p["balance"] for p in periodos]
        n_p = len(xs)
        sum_x  = sum(xs)
        sum_y  = sum(ys)
        sum_xy = sum(x * y for x, y in zip(xs, ys))
        sum_xx = sum(x * x for x in xs)
        denom  = (n_p * sum_xx - sum_x ** 2)
        slope  = ((n_p * sum_xy - sum_x * sum_y) / denom) if denom != 0 else 0
        tendencia = "creciente" if slope > 50 else ("decreciente" if slope < -50 else "estable")
    else:
        slope     = 0
        tendencia = "insuficiente"

    # ── Motor de reglas: mensajes e interpretaciones ─────────────────────────
    insights = []
    recomendaciones = []

    # Regla 1: salud financiera global
    if total_mrg >= 30:
        insights.append({"tipo": "positivo", "msg": f"Margen neto saludable: {total_mrg:.1f}%. El negocio genera valor real."})
    elif total_mrg >= 0:
        insights.append({"tipo": "neutro",   "msg": f"Margen neto ajustado: {total_mrg:.1f}%. Hay espacio de mejora."})
    else:
        insights.append({"tipo": "negativo", "msg": f"Balance negativo: {total_mrg:.1f}%. Los egresos superan los ingresos."})
        recomendaciones.append("Revisar los egresos más altos y evaluar reducción de costos operativos.")

    # Regla 2: tendencia
    if tendencia == "creciente":
        insights.append({"tipo": "positivo", "msg": "El balance muestra una tendencia creciente en el tiempo."})
    elif tendencia == "decreciente":
        insights.append({"tipo": "negativo", "msg": "El balance muestra una tendencia decreciente. Requiere atención."})
        recomendaciones.append("Identificar los periodos con mayor caída y analizar sus causas.")
    elif tendencia == "estable":
        insights.append({"tipo": "neutro",   "msg": "El balance es estable. Considerar estrategias de crecimiento."})

    # Regla 3: valores atípicos
    if atipicos:
        insights.append({"tipo": "alerta", "msg": f"Se detectaron {len(atipicos)} movimiento(s) atípico(s) que pueden distorsionar el análisis."})
        recomendaciones.append("Verificar los movimientos atípicos detectados para confirmar su validez.")

    # Regla 4: concentración de ingresos
    if len(periodos) > 1:
        max_ing_val = max_ing["ingresos"]
        if total_ing > 0 and (max_ing_val / total_ing) > 0.6:
            insights.append({"tipo": "alerta", "msg": f"El {(max_ing_val/total_ing*100):.0f}% de los ingresos se concentra en {max_ing['label']}. Alta dependencia."})
            recomendaciones.append("Diversificar las fuentes de ingreso para reducir la concentración de riesgo.")

    # Regla 5: variación fuerte en el último periodo
    if len(periodos) >= 2 and periodos[-1]["var_balance"] is not None:
        v = periodos[-1]["var_balance"]
        if v is not None and v < -25:
            insights.append({"tipo": "negativo", "msg": f"El último periodo muestra una caída del balance de {v:.1f}%."})
            recomendaciones.append(f"Analizar el período {periodos[-1]['label']} en detalle: revisar egresos extraordinarios.")
        elif v is not None and v > 50:
            insights.append({"tipo": "positivo", "msg": f"Excelente cierre: el último periodo creció {v:.1f}% vs el anterior."})

    # Regla 6: sin recomendaciones = todo OK
    if not recomendaciones:
        recomendaciones.append("Los indicadores están dentro de rangos saludables. Mantener el ritmo actual.")

    return {
        "empty":          False,
        "periodo":        periodo,
        "periodos":       periodos,
        "totales": {
            "ingresos": round(total_ing, 2),
            "egresos":  round(total_eg,  2),
            "balance":  round(total_bal, 2),
            "margen":   round(total_mrg, 2),
            "n":        n_total,
        },
        "mejor_periodo":      mejor,
        "peor_periodo":       peor,
        "max_ingresos":       max_ing,
        "max_egresos":        max_eg,
        "tendencia":          tendencia,
        "slope":              round(slope, 2),
        "atipicos":           atipicos,
        "insights":           insights,
        "recomendaciones":    recomendaciones,
        "generado":           datetime.now().isoformat(timespec="seconds"),
    }

# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.get("/analytics/data")
def analytics_data(
    periodo:    str           = Query("mensual", pattern="^(diario|mensual|anual)$"),
    start_date: Optional[str] = Query(None),
    end_date:   Optional[str] = Query(None),
    db:         Session       = Depends(get_db),
):
    q = db.query(Registro)
    if start_date:
        q = q.filter(Registro.fecha >= date.fromisoformat(start_date))
    if end_date:
        q = q.filter(Registro.fecha <= date.fromisoformat(end_date))
    registros = q.order_by(Registro.fecha).all()
    return compute_analytics(registros, periodo)


@router.get("/analytics/export/csv")
def export_csv(
    periodo:    str           = Query("mensual"),
    start_date: Optional[str] = Query(None),
    end_date:   Optional[str] = Query(None),
    db:         Session       = Depends(get_db),
):
    q = db.query(Registro)
    if start_date:
        q = q.filter(Registro.fecha >= date.fromisoformat(start_date))
    if end_date:
        q = q.filter(Registro.fecha <= date.fromisoformat(end_date))
    registros = q.order_by(Registro.fecha).all()
    data = compute_analytics(registros, periodo)

    buf = io.StringIO()
    w   = csv.writer(buf)

    # Header info
    w.writerow(["SEARI — Reporte de Análisis"])
    w.writerow(["Generado", data.get("generado", "")])
    w.writerow(["Periodo", periodo.capitalize()])
    w.writerow([])

    # Totales
    t = data["totales"]
    w.writerow(["RESUMEN GLOBAL"])
    w.writerow(["Total Ingresos", t["ingresos"]])
    w.writerow(["Total Egresos",  t["egresos"]])
    w.writerow(["Balance Neto",   t["balance"]])
    w.writerow(["Margen (%)",     t["margen"]])
    w.writerow(["Movimientos",    t["n"]])
    w.writerow([])

    # Detalle por periodo
    w.writerow(["DETALLE POR PERÍODO", "", "", "", "", "", "", ""])
    w.writerow(["Período", "Ingresos", "Egresos", "Balance", "Margen %",
                "Var. Ing %", "Var. Eg %", "Var. Bal %", "N° Mov."])
    for p in data.get("periodos", []):
        def fv(x): return f"{x:.2f}" if x is not None else "—"
        w.writerow([
            p["label"],
            p["ingresos"], p["egresos"], p["balance"],
            f"{p['margen']:.1f}%",
            fv(p["var_ingresos"]), fv(p["var_egresos"]), fv(p["var_balance"]),
            p["n"],
        ])
    w.writerow([])

    # Insights
    w.writerow(["INSIGHTS"])
    for ins in data.get("insights", []):
        w.writerow([ins["tipo"].upper(), ins["msg"]])
    w.writerow([])

    # Recomendaciones
    w.writerow(["RECOMENDACIONES"])
    for i, rec in enumerate(data.get("recomendaciones", []), 1):
        w.writerow([f"{i}.", rec])

    # Atípicos
    if data.get("atipicos"):
        w.writerow([])
        w.writerow(["MOVIMIENTOS ATÍPICOS"])
        w.writerow(["ID", "Concepto", "Valor", "Fecha", "Tipo", "Razón"])
        for a in data["atipicos"]:
            w.writerow([a["id"], a["concepto"], a["valor"], a["fecha"], a["tipo"], a["razon"]])

    buf.seek(0)
    filename = f"SEARI_Reporte_{periodo}_{date.today()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/analytics/export/excel")
def export_excel(
    periodo:    str           = Query("mensual"),
    start_date: Optional[str] = Query(None),
    end_date:   Optional[str] = Query(None),
    db:         Session       = Depends(get_db),
):
    q = db.query(Registro)
    if start_date:
        q = q.filter(Registro.fecha >= date.fromisoformat(start_date))
    if end_date:
        q = q.filter(Registro.fecha <= date.fromisoformat(end_date))
    registros = q.order_by(Registro.fecha).all()
    data = compute_analytics(registros, periodo)

    wb = Workbook()

    # ── Palette ──────────────────────────────────────────────────────────────
    C_BG_DARK   = "0D0D0F"
    C_SURFACE   = "141416"
    C_HEADER    = "1C1C20"
    C_BLUE      = "4F8EF7"
    C_GREEN     = "34D48A"
    C_RED       = "F74F4F"
    C_AMBER     = "F7A24F"
    C_TEXT      = "F0F0F2"
    C_MUTED     = "6B6B78"
    C_WHITE     = "FFFFFF"

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def font(color=C_TEXT, bold=False, size=10, name="Arial"):
        return Font(name=name, color=color, bold=bold, size=size)

    def border_bottom(color="3A3A42"):
        s = Side(style="thin", color=color)
        return Border(bottom=s)

    def thin_border(color="3A3A42"):
        s = Side(style="thin", color=color)
        return Border(top=s, bottom=s, left=s, right=s)

    def center():
        return Alignment(horizontal="center", vertical="center")

    def right():
        return Alignment(horizontal="right", vertical="center")

    FMT_CURRENCY = '$#,##0.00;($#,##0.00);"-"'
    FMT_PCT      = '0.0%'
    FMT_INT      = '#,##0'

    # ══ Sheet 1: Resumen Ejecutivo ════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_properties.tabColor = C_BLUE

    # Background whole sheet
    for row in ws1.iter_rows(min_row=1, max_row=60, min_col=1, max_col=10):
        for cell in row:
            cell.fill = fill(C_BG_DARK)

    # Title block
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "SEARI — REPORTE EJECUTIVO"
    ws1["A1"].font      = Font(name="Arial", color=C_BLUE,  bold=True, size=18)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A1"].fill      = fill(C_SURFACE)
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"Período: {periodo.capitalize()}   |   Generado: {data.get('generado','')}   |   Tendencia: {data.get('tendencia','').upper()}"
    ws1["A2"].font      = Font(name="Arial", color=C_MUTED, size=9)
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1["A2"].fill      = fill(C_SURFACE)
    ws1.row_dimensions[2].height = 18

    # ── KPI cards (row 4–8) ──────────────────────────────────────────────────
    t = data["totales"]
    kpis = [
        ("INGRESOS TOTALES", t["ingresos"], FMT_CURRENCY, C_GREEN),
        ("EGRESOS TOTALES",  t["egresos"],  FMT_CURRENCY, C_RED),
        ("BALANCE NETO",     t["balance"],  FMT_CURRENCY, C_BLUE),
        ("MARGEN NETO",      t["margen"] / 100, FMT_PCT,  C_AMBER),
        ("MOVIMIENTOS",      t["n"],        FMT_INT,      C_MUTED),
    ]

    col_map = [2, 4, 6, 8, 10]  # B, D, F, H, J
    for (label, value, fmt_str, color), col in zip(kpis, col_map):
        cl = get_column_letter(col)
        ws1.merge_cells(f"{cl}4:{cl}5")
        ws1[f"{cl}4"] = label
        ws1[f"{cl}4"].font      = Font(name="Arial", color=color, bold=True, size=8)
        ws1[f"{cl}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws1[f"{cl}4"].fill      = fill(C_HEADER)

        ws1.merge_cells(f"{cl}6:{cl}7")
        ws1[f"{cl}6"] = value
        ws1[f"{cl}6"].font         = Font(name="Arial", color=C_WHITE, bold=True, size=16)
        ws1[f"{cl}6"].alignment    = Alignment(horizontal="center", vertical="center")
        ws1[f"{cl}6"].fill         = fill(C_SURFACE)
        ws1[f"{cl}6"].number_format = fmt_str

    ws1.row_dimensions[4].height = 18
    ws1.row_dimensions[6].height = 30

    # ── Section: Detalle por Período ─────────────────────────────────────────
    start_row = 10
    ws1[f"A{start_row}"] = "ANÁLISIS POR PERÍODO"
    ws1[f"A{start_row}"].font = Font(name="Arial", color=C_BLUE, bold=True, size=11)
    ws1[f"A{start_row}"].fill = fill(C_BG_DARK)

    headers = ["Período", "Ingresos", "Egresos", "Balance", "Margen",
               "Var. Ing.", "Var. Eg.", "Var. Bal.", "Mov."]
    header_row = start_row + 1
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=c, value=h)
        cell.font      = Font(name="Arial", color=C_TEXT, bold=True, size=9)
        cell.fill      = fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    ws1.row_dimensions[header_row].height = 20

    for i, p in enumerate(data.get("periodos", [])):
        r = header_row + 1 + i
        row_fill = fill(C_SURFACE) if i % 2 == 0 else fill(C_HEADER)
        row_data = [
            (p["label"],       "General",      Alignment(horizontal="left",   vertical="center")),
            (p["ingresos"],    FMT_CURRENCY,   right()),
            (p["egresos"],     FMT_CURRENCY,   right()),
            (p["balance"],     FMT_CURRENCY,   right()),
            (p["margen"] / 100, FMT_PCT,       right()),
            (p["var_ingresos"] / 100 if p["var_ingresos"] is not None else None, FMT_PCT, right()),
            (p["var_egresos"]  / 100 if p["var_egresos"]  is not None else None, FMT_PCT, right()),
            (p["var_balance"]  / 100 if p["var_balance"]  is not None else None, FMT_PCT, right()),
            (p["n"],           FMT_INT,        center()),
        ]
        for c, (val, fmt_str, align) in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val if val is not None else "—")
            cell.font              = Font(name="Arial", color=C_TEXT, size=9)
            cell.fill              = row_fill
            cell.alignment         = align
            cell.number_format     = fmt_str
            cell.border            = border_bottom()
            # Color balance positive/negative
            if c == 4 and isinstance(val, (int, float)):
                cell.font = Font(name="Arial", color=C_GREEN if val >= 0 else C_RED, bold=True, size=9)
        ws1.row_dimensions[r].height = 16

    # Column widths Sheet1
    col_widths = [14, 14, 14, 14, 10, 10, 10, 10, 8]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Section: Insights ─────────────────────────────────────────────────────
    ins_start = header_row + len(data.get("periodos", [])) + 3
    ws1.merge_cells(f"A{ins_start}:I{ins_start}")
    ws1[f"A{ins_start}"] = "INTELIGENCIA BASADA EN REGLAS"
    ws1[f"A{ins_start}"].font = Font(name="Arial", color=C_BLUE, bold=True, size=11)
    ws1[f"A{ins_start}"].fill = fill(C_BG_DARK)

    tipo_color = {"positivo": C_GREEN, "negativo": C_RED, "alerta": C_AMBER, "neutro": C_MUTED}
    for i, ins in enumerate(data.get("insights", [])):
        r = ins_start + 1 + i
        ws1.merge_cells(f"B{r}:I{r}")
        ws1.cell(row=r, column=1, value=f"● {ins['tipo'].upper()}")\
            .font = Font(name="Arial", color=tipo_color.get(ins["tipo"], C_TEXT), bold=True, size=9)
        ws1.cell(row=r, column=1).fill = fill(C_SURFACE)
        ws1.cell(row=r, column=2, value=ins["msg"])\
            .font = Font(name="Arial", color=C_TEXT, size=9)
        ws1.cell(row=r, column=2).fill = fill(C_SURFACE)
        ws1.row_dimensions[r].height = 16

    rec_start = ins_start + len(data.get("insights", [])) + 2
    ws1.merge_cells(f"A{rec_start}:I{rec_start}")
    ws1[f"A{rec_start}"] = "RECOMENDACIONES"
    ws1[f"A{rec_start}"].font = Font(name="Arial", color=C_AMBER, bold=True, size=11)
    ws1[f"A{rec_start}"].fill = fill(C_BG_DARK)

    for i, rec in enumerate(data.get("recomendaciones", []), 1):
        r = rec_start + i
        ws1.merge_cells(f"A{r}:I{r}")
        ws1.cell(row=r, column=1, value=f"{i}. {rec}")\
            .font = Font(name="Arial", color=C_TEXT, size=9)
        ws1.cell(row=r, column=1).fill = fill(C_SURFACE)
        ws1.row_dimensions[r].height = 16

    # ══ Sheet 2: Movimientos Detallados ═══════════════════════════════════════
    ws2 = wb.create_sheet("Movimientos Detallados")
    ws2.sheet_properties.tabColor = C_GREEN

    for row in ws2.iter_rows(min_row=1, max_row=len(registros) + 5, min_col=1, max_col=6):
        for cell in row:
            cell.fill = fill(C_BG_DARK)

    ws2.merge_cells("A1:F1")
    ws2["A1"] = "MOVIMIENTOS DETALLADOS"
    ws2["A1"].font      = Font(name="Arial", color=C_GREEN, bold=True, size=13)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2["A1"].fill      = fill(C_SURFACE)
    ws2.row_dimensions[1].height = 28

    heads2 = ["ID", "Concepto", "Valor", "Fecha", "Tipo"]
    for c, h in enumerate(heads2, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font      = Font(name="Arial", color=C_TEXT, bold=True, size=9)
        cell.fill      = fill(C_HEADER)
        cell.alignment = center()
        cell.border    = thin_border()
    ws2.row_dimensions[2].height = 18

    for i, reg in enumerate(registros):
        r = 3 + i
        ing = reg.tipo.lower() == "ingreso"
        row_fill = fill(C_SURFACE) if i % 2 == 0 else fill(C_HEADER)
        vals = [reg.id, reg.concepto, float(reg.valor), str(reg.fecha), reg.tipo]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.fill  = row_fill
            cell.border = border_bottom()
            if c == 3:
                cell.number_format = FMT_CURRENCY
                cell.font = Font(name="Arial", color=C_GREEN if ing else C_RED, bold=True, size=9)
                cell.alignment = right()
            elif c == 5:
                cell.font = Font(name="Arial",
                                 color=C_GREEN if ing else C_RED, bold=True, size=9)
                cell.alignment = center()
            else:
                cell.font = Font(name="Arial", color=C_TEXT, size=9)
                cell.alignment = Alignment(vertical="center")
        ws2.row_dimensions[r].height = 15

    # Totales row
    tr = 3 + len(registros) + 1
    ws2.cell(row=tr, column=1, value="TOTALES").font = Font(name="Arial", color=C_BLUE, bold=True, size=9)
    ws2.cell(row=tr, column=1).fill = fill(C_HEADER)
    total_cell = ws2.cell(row=tr, column=3,
                          value=f'=SUM(C3:C{3+len(registros)-1})')
    total_cell.number_format = FMT_CURRENCY
    total_cell.font  = Font(name="Arial", color=C_BLUE, bold=True, size=10)
    total_cell.fill  = fill(C_HEADER)
    total_cell.alignment = right()
    ws2.row_dimensions[tr].height = 18

    col_widths2 = [8, 30, 16, 14, 12]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ══ Sheet 3: Valores Atípicos ═════════════════════════════════════════════
    ws3 = wb.create_sheet("Valores Atípicos")
    ws3.sheet_properties.tabColor = C_RED

    for row in ws3.iter_rows(min_row=1, max_row=30, min_col=1, max_col=7):
        for cell in row:
            cell.fill = fill(C_BG_DARK)

    ws3.merge_cells("A1:G1")
    ws3["A1"] = "DETECCIÓN DE VALORES ATÍPICOS (Método IQR)"
    ws3["A1"].font      = Font(name="Arial", color=C_RED, bold=True, size=13)
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3["A1"].fill      = fill(C_SURFACE)
    ws3.row_dimensions[1].height = 28

    if not data.get("atipicos"):
        ws3.merge_cells("A3:G3")
        ws3["A3"] = "✓ No se detectaron valores atípicos en el período seleccionado."
        ws3["A3"].font = Font(name="Arial", color=C_GREEN, size=10, bold=True)
        ws3["A3"].fill = fill(C_SURFACE)
    else:
        heads3 = ["ID", "Concepto", "Valor", "Fecha", "Tipo", "Razón"]
        for c, h in enumerate(heads3, 1):
            cell = ws3.cell(row=2, column=c, value=h)
            cell.font  = Font(name="Arial", color=C_TEXT, bold=True, size=9)
            cell.fill  = fill(C_HEADER)
            cell.alignment = center()
            cell.border = thin_border()

        for i, a in enumerate(data["atipicos"]):
            r = 3 + i
            for c, v in enumerate([a["id"], a["concepto"], a["valor"],
                                    a["fecha"], a["tipo"], a["razon"]], 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.font = Font(name="Arial", color=C_AMBER if c == 6 else C_TEXT, size=9)
                cell.fill = fill(C_SURFACE if i % 2 == 0 else C_HEADER)
                cell.border = border_bottom()
                if c == 3:
                    cell.number_format = FMT_CURRENCY
                    cell.alignment = right()

        col_widths3 = [8, 28, 16, 14, 12, 18]
        for i, w in enumerate(col_widths3, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SEARI_Reporte_{periodo}_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/analytics/export/pdf")
def export_pdf(
    periodo:    str           = Query("mensual"),
    start_date: Optional[str] = Query(None),
    end_date:   Optional[str] = Query(None),
    db:         Session       = Depends(get_db),
):
    q = db.query(Registro)
    if start_date:
        q = q.filter(Registro.fecha >= date.fromisoformat(start_date))
    if end_date:
        q = q.filter(Registro.fecha <= date.fromisoformat(end_date))
    registros = q.order_by(Registro.fecha).all()
    data = compute_analytics(registros, periodo)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=16*mm, bottomMargin=16*mm,
    )

    # ── Colors ────────────────────────────────────────────────────────────────
    BG     = colors.HexColor("#0D0D0F")
    SURF   = colors.HexColor("#141416")
    SURF2  = colors.HexColor("#1C1C20")
    BORDER = colors.HexColor("#3A3A42")
    TEXT   = colors.HexColor("#F0F0F2")
    MUTED  = colors.HexColor("#6B6B78")
    BLUE   = colors.HexColor("#4F8EF7")
    GREEN  = colors.HexColor("#34D48A")
    RED    = colors.HexColor("#F74F4F")
    AMBER  = colors.HexColor("#F7A24F")

    # ── Styles ────────────────────────────────────────────────────────────────
    def style(name, parent=None, **kw):
        defaults = dict(fontName="Helvetica", fontSize=9, textColor=TEXT,
                        backColor=BG, leading=13)
        defaults.update(kw)
        return ParagraphStyle(name, **defaults)

    S_TITLE    = style("title",   fontSize=20, textColor=BLUE,  alignment=TA_CENTER,
                       fontName="Helvetica-Bold", spaceAfter=2)
    S_SUBTITLE = style("sub",     fontSize=8,  textColor=MUTED, alignment=TA_CENTER)
    S_SECTION  = style("section", fontSize=11, textColor=BLUE,  fontName="Helvetica-Bold",
                       spaceBefore=10, spaceAfter=4)
    S_BODY     = style("body",    fontSize=8,  textColor=TEXT)
    S_INSIGHT  = style("insight", fontSize=8,  textColor=TEXT, leftIndent=8)
    S_REC      = style("rec",     fontSize=8,  textColor=AMBER, leftIndent=8)

    def hr(): return HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=6)

    story = []

    # ── Cover ─────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 8*mm))
    story.append(Paragraph("SEARI", S_TITLE))
    story.append(Paragraph("Reporte Ejecutivo de Análisis Financiero", S_SUBTITLE))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"Período: {periodo.capitalize()}  ·  Generado: {data.get('generado','')}  ·  Tendencia: {data.get('tendencia','').upper()}",
        S_SUBTITLE
    ))
    story.append(Spacer(1, 6*mm))
    story.append(hr())

    # ── KPI table ─────────────────────────────────────────────────────────────
    t = data["totales"]
    kpi_data = [
        ["INGRESOS",        "EGRESOS",       "BALANCE",       "MARGEN",           "MOVIMIENTOS"],
        [fmt(t["ingresos"]), fmt(t["egresos"]), fmt(t["balance"]), f"{t['margen']:.1f}%", str(t["n"])],
    ]
    kpi_table = Table(kpi_data, colWidths=[35*mm]*5)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0),  SURF2),
        ("BACKGROUND",  (0,1), (-1,1),  SURF),
        ("TEXTCOLOR",   (0,0), (0,-1),  GREEN),
        ("TEXTCOLOR",   (1,0), (1,-1),  RED),
        ("TEXTCOLOR",   (2,0), (2,-1),  BLUE),
        ("TEXTCOLOR",   (3,0), (3,-1),  AMBER),
        ("TEXTCOLOR",   (4,0), (4,-1),  MUTED),
        ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTNAME",    (0,1), (-1,1),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0),  7),
        ("FONTSIZE",    (0,1), (-1,1),  13),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [SURF2, SURF]),
        ("GRID",        (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 5*mm))

    # ── Detalle por periodo ───────────────────────────────────────────────────
    story.append(Paragraph("ANÁLISIS POR PERÍODO", S_SECTION))

    def fv(x): return f"{x:.1f}%" if x is not None else "—"
    table_data = [["Período", "Ingresos", "Egresos", "Balance", "Margen",
                   "Var. Ing.", "Var. Bal.", "Mov."]]
    for p in data.get("periodos", []):
        table_data.append([
            p["label"],
            fmt(p["ingresos"]), fmt(p["egresos"]), fmt(p["balance"]),
            f"{p['margen']:.1f}%",
            fv(p["var_ingresos"]), fv(p["var_balance"]),
            str(p["n"]),
        ])

    col_w = [22*mm, 22*mm, 22*mm, 22*mm, 16*mm, 16*mm, 16*mm, 12*mm]
    dt_table = Table(table_data, colWidths=col_w, repeatRows=1)
    ts = [
        ("BACKGROUND",   (0,0), (-1,0),  SURF2),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 7),
        ("TEXTCOLOR",    (0,0), (-1,0),  TEXT),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (0,0), (0,-1),  "LEFT"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("GRID",         (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]
    for i, p in enumerate(data.get("periodos", []), 1):
        bg = SURF if i % 2 == 0 else SURF2
        ts.append(("BACKGROUND", (0,i), (-1,i), bg))
        ts.append(("TEXTCOLOR",  (0,i), (-1,i), TEXT))
        # Balance color
        bal_color = GREEN if p["balance"] >= 0 else RED
        ts.append(("TEXTCOLOR",  (3,i), (3,i),  bal_color))
        ts.append(("FONTNAME",   (3,i), (3,i),  "Helvetica-Bold"))
    dt_table.setStyle(TableStyle(ts))
    story.append(dt_table)
    story.append(Spacer(1, 5*mm))

    # ── Insights ──────────────────────────────────────────────────────────────
    story.append(hr())
    story.append(Paragraph("INTELIGENCIA BASADA EN REGLAS", S_SECTION))
    tipo_map = {"positivo": ("✓", GREEN), "negativo": ("✗", RED),
                "alerta": ("⚠", AMBER), "neutro": ("●", MUTED)}
    for ins in data.get("insights", []):
        icon, color = tipo_map.get(ins["tipo"], ("•", TEXT))
        ps = ParagraphStyle(f"ins_{ins['tipo']}", parent=S_INSIGHT,
                            textColor=color, fontSize=8, leading=12)
        story.append(Paragraph(f"{icon}  <b>{ins['tipo'].upper()}</b>  {ins['msg']}", ps))
        story.append(Spacer(1, 1*mm))

    # ── Recomendaciones ───────────────────────────────────────────────────────
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph("RECOMENDACIONES", S_SECTION))
    for i, rec in enumerate(data.get("recomendaciones", []), 1):
        story.append(Paragraph(f"{i}.  {rec}", S_REC))
        story.append(Spacer(1, 1*mm))

    # ── Atípicos ──────────────────────────────────────────────────────────────
    if data.get("atipicos"):
        story.append(Spacer(1, 3*mm))
        story.append(hr())
        story.append(Paragraph("MOVIMIENTOS ATÍPICOS DETECTADOS", S_SECTION))
        at_data = [["ID", "Concepto", "Valor", "Fecha", "Tipo", "Razón"]]
        for a in data["atipicos"]:
            at_data.append([str(a["id"]), a["concepto"], fmt(a["valor"]),
                            a["fecha"], a["tipo"], a["razon"]])
        at_table = Table(at_data, colWidths=[10*mm, 45*mm, 22*mm, 22*mm, 16*mm, 28*mm])
        at_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0),  SURF2),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 7),
            ("TEXTCOLOR",    (0,0), (-1,0),  TEXT),
            ("TEXTCOLOR",    (0,1), (-1,-1), TEXT),
            ("TEXTCOLOR",    (5,1), (5,-1),  AMBER),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("ALIGN",        (1,0), (1,-1),  "LEFT"),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("GRID",         (0,0), (-1,-1), 0.3, BORDER),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [SURF, SURF2]),
            ("TOPPADDING",   (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ]))
        story.append(at_table)

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 8*mm))
    story.append(hr())
    story.append(Paragraph(
        f"SEARI · Sistema de Análisis e Inteligencia Financiera · {date.today().isoformat()}",
        style("footer", fontSize=7, textColor=MUTED, alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"SEARI_Reporte_{periodo}_{date.today()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )